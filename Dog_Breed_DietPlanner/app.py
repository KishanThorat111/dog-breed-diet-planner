from flask import Flask, render_template, request, redirect, url_for, session,send_file
import os
import numpy as np
from PIL import Image
import tensorflow as tf
from tensorflow.keras.applications.mobilenet_v2 import preprocess_input, decode_predictions
import openpyxl
from diet_data import diet_plans, get_size_category
import datetime

app = Flask(__name__)
app.secret_key = "secretkey"
app.config['UPLOAD_FOLDER'] = 'static/uploads'

# Excel file
EXCEL_FILE = 'users.xlsx'
ADMIN_EMAIL = "admin@gmail.com"
ADMIN_PASSWORD = "admin123"

# Load pretrained MobileNetV2 model
model = tf.keras.applications.MobileNetV2(weights='imagenet')

# ==============================
# Excel User Functions
# ==============================

def register_user(username, email, password):
    # Create Excel file if it does not exist
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Username", "Email", "Password"])
        wb.save(EXCEL_FILE)
    
def check_user(email, password):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1] == email and row[2] == password:
            return row[0]  # return username
    return None

# ==============================
# Breed Prediction
# ==============================

def predict_breed(img_path):
    img = Image.open(img_path).resize((224, 224))
    img_array = np.array(img)
    img_array = preprocess_input(img_array)
    img_array = np.expand_dims(img_array, axis=0)

    predictions = model.predict(img_array)
    decoded = decode_predictions(predictions, top=1)[0][0]

    breed_name = decoded[1]
    confidence = round(decoded[2] * 100, 2)

    return breed_name, confidence

# ==============================
# Routes
# ==============================

@app.route('/')
def home_redirect():
    return redirect(url_for('login'))

# ------------------------------
# Register
# ------------------------------
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')

        # Check if any field is missing
        if not username or not email or not password:
            return render_template('register.html', 
                                   message="All fields are required!")

        register_user(username, email, password)

        return render_template('login.html', 
                               message="Registration Successful! Please Login")

    return render_template('register.html')
# ------------------------------
# Login
# ------------------------------
@app.route('/login', methods=['GET', 'POST'])
def login():

    if request.method == 'POST':

        email = request.form['email']
        password = request.form['password']

        # Admin Login
        if email == ADMIN_EMAIL and password == ADMIN_PASSWORD:
            session['admin'] = True
            return redirect(url_for('admin_dashboard'))

        # Normal User Login
        user = check_user(email, password)

        if user:
            session['user'] = user
            return redirect(url_for('index'))

        else:
            return render_template('login.html', message="Invalid Credentials")

    return render_template('login.html')
# ------------------------------
# Logout
# ------------------------------
@app.route('/logout')
def logout():
    session.pop('user', None)
    return redirect(url_for('login'))

# ------------------------------
# Main App (Protected)
# ------------------------------
@app.route('/index', methods=['GET', 'POST'])
def index():
    if 'user' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        file = request.files['image']

        if file:
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(filepath)

            breed, confidence = predict_breed(filepath)

            size_category = get_size_category(breed)
            diet = diet_plans[size_category]
            
            # YouTube training link
            youtube_link = f"https://www.youtube.com/results?search_query={breed}+dog+training+exercise"

            # Amazon link for food
            food_name = diet["food"]
            amazon_link = "https://www.amazon.in/s?k=" + food_name.replace(" ", "+")

            # Store for report download
            session['breed'] = breed
            session['confidence'] = confidence
            session['size'] = size_category
            session['diet'] = diet

            return render_template(
                'result.html',
                image=filepath,
                breed=breed,
                confidence=confidence,
                diet=diet,
                size=size_category,
                youtube_link=youtube_link,
                amazon_link=amazon_link,
                user=session['user']
            )
    return render_template('index.html', user=session['user'])

@app.route('/download_report')
def download_report():

    if 'user' not in session:
        return redirect(url_for('login'))

    breed = session.get('breed')
    confidence = session.get('confidence')
    size = session.get('size')
    diet = session.get('diet')

    report_text = f"""
Dog Identification & Diet Report
---------------------------------

Breed: {breed}
Confidence: {confidence} %

Size Category: {size}

Recommended Diet Plan
---------------------
Main Food: {diet['food']}
Meals Per Day: {diet['meals']}
Extras: {diet['extras']}

Generated on: {datetime.datetime.now()}
"""

    file_path = "dog_report.txt"

    with open(file_path, "w") as f:
        f.write(report_text)

    return send_file(file_path, as_attachment=True)

@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():

    if request.method == 'POST':

        email = request.form['email']
        password = request.form['password']

        if email == "admin@gmail.com" and password == "admin123":
            session['admin'] = True
            return redirect(url_for('admin_dashboard'))

        else:
            return render_template("admin_login.html", message="Invalid Admin Credentials")

    return render_template("admin_login.html")

@app.route('/admin')
def admin_dashboard():

    if 'admin' not in session:
        return redirect(url_for('admin_login'))

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    users = []

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        users.append({
            "id": index,
            "username": row[0],
            "email": row[1],
            "password": row[2]
        })

    return render_template("admin.html", users=users)

@app.route('/delete_user/<int:user_id>')
def delete_user(user_id):

    if 'admin' not in session:
        return redirect(url_for('admin_login'))

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    sheet.delete_rows(user_id + 2)

    wb.save(EXCEL_FILE)

    return redirect(url_for('admin_dashboard'))

@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
def edit_user(user_id):

    if 'admin' not in session:
        return redirect(url_for('admin_login'))

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    row_number = user_id + 2

    if request.method == 'POST':

        username = request.form['username']
        email = request.form['email']
        password = request.form['password']

        sheet.cell(row=row_number, column=1).value = username
        sheet.cell(row=row_number, column=2).value = email
        sheet.cell(row=row_number, column=3).value = password

        wb.save(EXCEL_FILE)

        return redirect(url_for('admin_dashboard'))

    user = {
        "username": sheet.cell(row=row_number, column=1).value,
        "email": sheet.cell(row=row_number, column=2).value,
        "password": sheet.cell(row=row_number, column=3).value
    }

    return render_template("edit_user.html", user=user)

@app.route('/admin_logout')
def admin_logout():
    session.pop('admin', None)
    return redirect(url_for('login'))

@app.route('/about')
def about():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('about.html', user=session['user'])


@app.route('/services')
def services():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('services.html', user=session['user'])


@app.route('/vets')
def vets():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('vets.html', user=session['user'])


# ==============================

if __name__ == '__main__':
    app.run(debug=True)
